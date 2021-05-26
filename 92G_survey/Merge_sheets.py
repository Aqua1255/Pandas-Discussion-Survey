# importing openpyxl module
import openpyxl as xl
from Excel_helper import append_df_to_excel
from survey_ghost import gender as gender_ghost
from survey_Courage import gender as gender_courage
from survey_Lancer import gender as gender_lancer
from survey_ghost import age as age_ghost
from survey_Courage import age as age_courage
from survey_Lancer import age as age_lancer
from survey_ghost import rank as rank_ghost
from survey_Lancer import rank as rank_lancer
from survey_Courage import rank as rank_courage
from

# opening the source excel files
lancer = '../../Desktop/Work/92G_survey_lancer.xlsx'
ghost = '../../Desktop/Work/92G_survey_ghost.xlsx'
courage = '../../Desktop/Work/92G_survey_courage.xlsx'

# open target excel files
total = '../../Desktop/Work/92G_survey_total.xlsx'

lancer_wb = xl.load_workbook(lancer)
ghost_wb = xl.load_workbook(ghost)
courage_wb = xl.load_workbook(courage)

gender_sheets = [gender_ghost, gender_courage, gender_lancer]
age_sheets = [gender_ghost, gender_courage, gender_lancer]
rank_sheets =[rank_ghost, rank_courage, rank_lancer]
personality_sheets =
friends_sheets =
off_base_sheets =
rec_activities_sheets =
reenlist_sheets =
squad_sheets =
company_sheets =
bn_sheets =
satisfaction_sheets =
command_support_sheets =
concern_sheets =
concern_sheets =
covax_sheets =
suggestion_sheets =
sharp_sheets =
leadership_sheets =
self_harm_sheets =
meal_quality_sheets =
live_location_sheets =
education_sheets =
games_hrs_sheets =
games_habits_sheets =
bills_sheets =



for each in gender_sheets:
    append_df_to_excel(total, each, sheet_name='gender', index=True)

# for sheet in lancer_wb:
#     gender = lancer_wb.worksheets[0
#     agec = lancer_wb.worksheets[1]

